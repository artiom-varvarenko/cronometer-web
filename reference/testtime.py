import tkinter as tk
from tkinter import messagebox
import time
import threading
import winsound
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import os
import math

class TestTimeApp:
    def __init__(self, root):
        self.root = root
        self.root.title("TestTIME_v0.2")
        self.root.geometry("720x650")
        self.root.resizable(True, True)  # Проблема 1: Возможность изменить размер окна
        self.root.configure(bg='#f0f0f0')

        # Переменные
        self.surname = tk.StringVar()
        self.intervals = [2, 3, 4, 5]  # Длительности интервалов в секундах
        self.interval_vars = [tk.StringVar(value=str(i)) for i in self.intervals]
        self.interval_counts = [0, 0, 0, 0]  # Счетчики использования каждого интервала
        self.results = []  # Результаты тестов
        self.user_start_time = 0
        self.is_user_timing = False
        self.is_playing_interval = False
        self.current_interval_type = -1
        self.total_tests = 0
        self.action_button = None  # Будет создана позже

        # Создание интерфейса
        self.create_initial_screen()

    def create_initial_screen(self):
        """Создание начального экрана"""
        # Заголовок
        self.title_label = tk.Label(self.root,
                                    text="Для начала работы введите фамилию и подтвердите",
                                    font=("Arial", 16, "italic"), pady=20, bg='#f0f0f0')
        self.title_label.pack()

        # Фрейм для фамилии и кнопки подтвердить
        top_frame = tk.Frame(self.root, bg='#f0f0f0')
        top_frame.pack(pady=10)

        # Поле ввода фамилии
        self.surname_entry = tk.Entry(top_frame, textvariable=self.surname,
                                      font=("Arial", 14), width=20, relief=tk.SOLID, borderwidth=1)
        self.surname_entry.pack(side=tk.LEFT, padx=10)

        self.confirm_btn = tk.Button(top_frame, text="Подтвердить",
                                     command=self.confirm_surname,
                                     font=("Arial", 12), padx=20, pady=5,
                                     relief=tk.RAISED, borderwidth=2)
        self.confirm_btn.pack(side=tk.LEFT, padx=10)

        # Центральная часть
        self.center_frame = tk.Frame(self.root, bg='#f0f0f0')
        self.center_frame.pack(pady=20)

        # Левая часть - кнопки интервалов
        left_frame = tk.Frame(self.center_frame, bg='#f0f0f0')
        left_frame.pack(side=tk.LEFT, padx=20)

        self.interval_buttons = []
        for i in range(4):
            btn = tk.Button(left_frame, text=f"Запустить {i+1}{'ый' if i == 0 or i == 3 else 'ой' if i == 1 else 'ий'} интервал",
                            command=lambda idx=i: self.play_interval(idx),
                            state=tk.DISABLED,
                            font=("Arial", 12), width=25, pady=8,
                            relief=tk.RAISED, borderwidth=2, bg='#e0e0e0')
            btn.pack(pady=5)
            self.interval_buttons.append(btn)

        # Правая часть - область Пуск/Стоп
        self.right_frame = tk.Frame(self.center_frame, relief=tk.SOLID, borderwidth=2,
                                    bg='white', width=200, height=250)
        self.right_frame.pack(side=tk.LEFT, padx=50)
        self.right_frame.pack_propagate(False)

        self.action_label = tk.Label(self.right_frame, text="Пуск",
                                     font=("Arial", 16, "underline"), pady=50, bg='white')
        self.action_label.pack()

        # Статус
        self.status_label = tk.Label(self.root, text="",
                                     font=("Arial", 14, "italic"), pady=10, bg='#f0f0f0')
        self.status_label.pack()

        # Нижние кнопки
        bottom_frame = tk.Frame(self.root, bg='#f0f0f0')
        bottom_frame.pack(side=tk.BOTTOM, pady=30)

        self.table_btn = tk.Button(bottom_frame, text="Отобразить таблицу результатов",
                                   command=self.show_results,
                                   state=tk.DISABLED,
                                   font=("Arial", 12), padx=20, pady=5,
                                   relief=tk.RAISED, borderwidth=2, bg='#e0e0e0')
        self.table_btn.pack(pady=5)

        self.intervals_btn = tk.Button(bottom_frame, text="Отобразить интервалы",
                                       command=self.show_intervals_config,
                                       font=("Arial", 12), padx=20, pady=5,
                                       relief=tk.RAISED, borderwidth=2)
        self.intervals_btn.pack(pady=5)

    def show_intervals_config(self):
        """Показать окно настройки интервалов"""
        config_window = tk.Toplevel(self.root)
        config_window.title("Настройка интервалов")
        config_window.geometry("400x350")
        config_window.resizable(False, False)

        tk.Label(config_window, text="Длительность интервалов (секунды):",
                 font=("Arial", 14), pady=20).pack()

        entries = []
        for i in range(4):
            frame = tk.Frame(config_window)
            frame.pack(pady=10)

            tk.Label(frame, text=f"{i+1}-{'й' if i == 0 or i == 3 else 'й' if i == 1 else 'й'} интервал:",
                     font=("Arial", 12), width=15).pack(side=tk.LEFT)

            entry = tk.Entry(frame, textvariable=self.interval_vars[i],
                             font=("Arial", 12), width=10)
            entry.pack(side=tk.LEFT)
            entries.append(entry)

            tk.Label(frame, text="сек.", font=("Arial", 12)).pack(side=tk.LEFT, padx=5)

        def save_intervals():
            try:
                for i in range(4):
                    value = float(self.interval_vars[i].get())
                    if value <= 0:
                        raise ValueError
                    self.intervals[i] = value
                config_window.destroy()
            except ValueError:
                messagebox.showerror("Ошибка", "Введите корректные положительные числа")

        tk.Button(config_window, text="Сохранить", command=save_intervals,
                  font=("Arial", 12), padx=30, pady=8).pack(pady=30)

    def confirm_surname(self):
        """Подтвердить фамилию и начать тест"""
        if not self.surname.get().strip():
            messagebox.showerror("Ошибка", "Введите фамилию")
            return

        # Обновить интерфейс
        self.surname_entry.config(state=tk.DISABLED)
        self.confirm_btn.config(state=tk.DISABLED, relief=tk.SUNKEN)

        # Активировать кнопки интервалов
        for btn in self.interval_buttons:
            btn.config(state=tk.NORMAL, bg='white')

        # Создать кнопку Пуск в правой панели
        self.create_action_button()

        # Проблема 3: Активировать кнопку таблицы результатов
        self.table_btn.config(state=tk.NORMAL)

        self.status_label.config(text="Выберите интервал для тестирования")

    def create_action_button(self):
        """Создать кнопку Пуск/Стоп"""
        # Удалить все виджеты из правой панели
        for widget in self.right_frame.winfo_children():
            widget.destroy()

        # Создать кнопку, которая выглядит как label
        self.action_button = tk.Button(self.right_frame, text="Пуск",
                                       command=self.toggle_timer,
                                       font=("Arial", 16, "underline"),
                                       width=12, height=4,
                                       state=tk.DISABLED,
                                       relief=tk.FLAT,
                                       bg='white',
                                       activebackground='#f0f0f0',
                                       bd=0)
        self.action_button.pack(expand=True, fill=tk.BOTH)

        # Сохраним ссылку на action_label для обратной совместимости
        self.action_label = self.action_button

    def play_interval(self, interval_index):
        """Воспроизвести выбранный интервал"""
        if self.is_playing_interval or self.is_user_timing:
            return

        # Проверить, не использовался ли интервал 5 раз
        if self.interval_counts[interval_index] >= 5:
            interval_names = ["1-ый", "2-ой", "3-ий", "4-ый"]
            messagebox.showwarning("Предупреждение",
                                   f"{interval_names[interval_index]} интервал уже использован 5 раз")
            return

        self.current_interval_type = interval_index
        self.is_playing_interval = True

        # Отключить все кнопки интервалов
        for btn in self.interval_buttons:
            btn.config(state=tk.DISABLED)

        # Отключить кнопку Пуск
        self.action_button.config(state=tk.DISABLED)

        interval_duration = self.intervals[interval_index]
        interval_names = ["1-го", "2-го", "3-го", "4-го"]
        self.status_label.config(text=f"Воспроизведение {interval_names[interval_index]} интервала...")

        # Воспроизвести интервал в отдельном потоке
        def play():
            # Начальный сигнал
            self.play_beep()
            time.sleep(interval_duration)
            # Конечный сигнал
            self.play_beep()

            # Обновить интерфейс
            self.root.after(0, self.interval_played)

        threading.Thread(target=play, daemon=True).start()

    def interval_played(self):
        """Интервал воспроизведен, активировать кнопку Пуск"""
        self.is_playing_interval = False
        self.action_button.config(state=tk.NORMAL)
        self.status_label.config(text="Теперь попробуйте воспроизвести этот интервал")

    def toggle_timer(self):
        """Переключить таймер Пуск/Стоп"""
        if not self.is_user_timing:
            # Начать отсчет
            self.is_user_timing = True
            self.user_start_time = time.time()
            self.action_button.config(text="Стоп")
            self.play_beep()  # Звуковой сигнал при нажатии Пуск
            self.status_label.config(text="Отсчет времени запущен")
        else:
            # Остановить отсчет
            self.is_user_timing = False
            elapsed_time = time.time() - self.user_start_time
            self.play_beep()  # Звуковой сигнал при нажатии Стоп

            # Сохранить результат
            self.results.append({
                'interval_type': self.current_interval_type,
                'user_time': elapsed_time,
                'target_time': self.intervals[self.current_interval_type]
            })

            # Увеличить счетчики
            self.interval_counts[self.current_interval_type] += 1
            self.total_tests += 1

            # Сбросить кнопку
            self.action_button.config(text="Пуск", state=tk.DISABLED)

            # Обновить статус
            self.status_label.config(text=f"Результат записан. (Тест {self.total_tests}/20)")

            # Проверить завершение
            if self.total_tests >= 20:
                self.finish_tests()
            else:
                # Активировать кнопки интервалов для следующего теста через небольшую задержку
                self.root.after(1000, self.enable_interval_buttons)

    def enable_interval_buttons(self):
        """Активировать доступные кнопки интервалов"""
        has_available = False
        for i, btn in enumerate(self.interval_buttons):
            if self.interval_counts[i] < 5:
                btn.config(state=tk.NORMAL, bg='white')
                has_available = True
            else:
                btn.config(state=tk.DISABLED, bg='#e0e0e0')

        # Сбросить состояние кнопки Пуск
        self.action_button.config(text="Пуск")

        # Обновить статус
        if has_available:
            available_intervals = [i+1 for i in range(4) if self.interval_counts[i] < 5]
            interval_texts = []
            for num in available_intervals:
                if num == 1 or num == 4:
                    interval_texts.append(f"{num}-ый")
                elif num == 2:
                    interval_texts.append(f"{num}-ой")
                else:
                    interval_texts.append(f"{num}-ий")
            self.status_label.config(text=f"Выберите следующий интервал. Доступны: {', '.join(interval_texts)}")
        else:
            # Это не должно произойти, так как у нас всего 20 тестов
            self.finish_tests()

    def finish_tests(self):
        """Завершить тестирование"""
        self.status_label.config(text="Тест завершен! Все 20 интервалов пройдены.")
        self.table_btn.config(state=tk.NORMAL)

        # Отключить все кнопки интервалов
        for btn in self.interval_buttons:
            btn.config(state=tk.DISABLED)

        # Изменить кнопку на "Скрыть интервалы"
        self.action_button.config(text="Скрыть интервалы", state=tk.NORMAL, command=lambda: None)

        messagebox.showinfo("Тест завершен",
                            "Вы прошли все 20 интервалов!\n" +
                            "Нажмите 'Отобразить таблицу результатов' для просмотра.")

    def play_beep(self):
        """Воспроизвести звуковой сигнал"""
        try:
            winsound.Beep(1000, 100)  # 1000 Hz, 100 мс
        except:
            pass

    def calculate_sigma(self, values, mean):
        """Рассчитать стандартное отклонение (несмещенная оценка)"""
        if len(values) <= 1:
            return 0
        variance = sum((x - mean) ** 2 for x in values) / (len(values) - 1)
        return math.sqrt(variance)

    def years_to_age_string(self, years):
        """Конвертировать годы в строку возраста (годы, месяцы, дни)"""
        total_days = years * 365.25  # Учитываем високосные годы

        # Извлекаем годы
        full_years = int(years)
        remaining_days = total_days - (full_years * 365.25)

        # Извлекаем месяцы (средняя длина месяца = 30.44 дня)
        full_months = int(remaining_days / 30.44)
        remaining_days -= full_months * 30.44

        # Округляем дни
        full_days = round(remaining_days)

        # Корректируем если дни >= 30
        if full_days >= 30:
            full_months += 1
            full_days -= 30

        # Корректируем если месяцы >= 12
        if full_months >= 12:
            full_years += 1
            full_months -= 12

        # Формируем строку
        parts = []
        if full_years > 0:
            parts.append(f"{full_years}г")
        if full_months > 0:
            parts.append(f"{full_months}м")
        if full_days > 0 or len(parts) == 0:  # Всегда показываем дни, если нет лет и месяцев
            parts.append(f"{full_days}д")

        return f"({', '.join(parts)})"

    def show_results(self):
        """Показать результаты в Excel"""
        if len(self.results) == 0:
            messagebox.showinfo("Информация", "Пока нет результатов для отображения")
            return

        # Создать Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Результаты"

        # Группировать результаты по типу интервала
        interval_results = [[] for _ in range(4)]
        for result in self.results:
            if result['interval_type'] >= 0:  # Проверка на валидность
                interval_results[result['interval_type']].append(result['user_time'])

        # Заголовки столбцов
        headers = ["Интервал", "Попытка 1", "Попытка 2", "Попытка 3", "Попытка 4", "Попытка 5", "Сумма", "Тау (т)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Записать результаты и подсчеты
        interval_names = ["2 секунды", "3 секунды", "4 секунды", "5 секунд"]
        row_data = []
        tau_values = []  # Для расчета квадратичного отклонения

        for i in range(4):
            row = i + 2
            # Название интервала
            ws.cell(row=row, column=1, value=interval_names[i])

            # Результаты в следующих 5 столбцах
            row_times = []
            for j in range(5):
                if j < len(interval_results[i]):
                    time_value = round(interval_results[i][j], 3)
                    ws.cell(row=row, column=j+2, value=time_value)
                    row_times.append(interval_results[i][j])
                else:
                    ws.cell(row=row, column=j+2, value="-")

            # Подсчет суммы
            if row_times:
                suma = sum(row_times)
                ws.cell(row=row, column=7, value=round(suma, 3))

                # Подсчет Тау
                tau = suma / (self.intervals[i] * 5)
                ws.cell(row=row, column=8, value=round(tau, 3))
                tau_values.append(tau)  # Сохраняем для расчета отклонения

                row_data.append({
                    'interval': i,
                    'times': row_times,
                    'suma': suma,
                    'tau': tau
                })

        # Добавить строку "Узагальнено"
        if row_data:
            row = 6
            ws.cell(row=row, column=1, value="Общее")
            for col in range(2, 7):
                ws.cell(row=row, column=col, value="-")

            # Общая сумма
            total_suma = sum(data['suma'] for data in row_data)
            ws.cell(row=row, column=7, value=round(total_suma, 3))

            # Общий Тау (среднее значение тау)
            avg_tau = sum(data['tau'] for data in row_data) / len(row_data)
            cell = ws.cell(row=row, column=8, value=round(avg_tau, 3))
            cell.font = Font(bold=True)

            # Добавить строку "Квадр. откл."
            row = 7
            ws.cell(row=row, column=1, value="Квадр. откл.")
            for col in range(2, 8):
                ws.cell(row=row, column=col, value="-")

            # Рассчитать среднее квадратичное отклонение из тау
            if tau_values:
                sigma = self.calculate_sigma(tau_values, avg_tau)
                ws.cell(row=row, column=8, value=round(sigma, 3))

            # Добавить таблицу циклов
            # Пропустить 2 строки
            row = 10
            ws.cell(row=row, column=4, value="Циклы (возраст)")
            ws.cell(row=row, column=4).font = Font(bold=True)
            ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')

            # Объединить ячейки для заголовка
            ws.merge_cells(f'D{row}:E{row}')

            # Заголовки таблицы циклов
            row = 12
            ws.cell(row=row, column=4, value="Цикл")
            ws.cell(row=row, column=5, value="Значения")
            ws.cell(row=row, column=4).font = Font(bold=True)
            ws.cell(row=row, column=5).font = Font(bold=True)

            # Добавить границы для заголовков
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            ws.cell(row=row, column=4).border = thin_border
            ws.cell(row=row, column=5).border = thin_border

            # Генерировать циклы от 0.25 до 12 с шагом 0.25
            cycles = []
            multiplier = 0.25
            while multiplier <= 12:
                cycle_name = f"C{multiplier:.2f}".rstrip('0').rstrip('.')
                cycles.append((cycle_name, multiplier))
                multiplier += 0.25

            # Рассчитать и записать значения циклов
            for idx, (cycle_name, multiplier) in enumerate(cycles):
                row = 13 + idx
                ws.cell(row=row, column=4, value=cycle_name).border = thin_border

                # Формула: Общее Тау (жирным шрифтом) * 8.5 * множитель цикла
                cycle_value = avg_tau * 8.5 * multiplier
                age_string = self.years_to_age_string(cycle_value)
                cell_value = f"{round(cycle_value, 3)}\n{age_string}"
                cell = ws.cell(row=row, column=5, value=cell_value)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='center')

        # Настроить ширину столбцов
        ws.column_dimensions['A'].width = 15
        for col in 'BCDEFGH':
            if col == 'E':
                ws.column_dimensions[col].width = 15  # Шире для значений с возрастом
            else:
                ws.column_dimensions[col].width = 10

        # Сохранить и открыть файл
        filename = f"{self.surname.get()}.xlsx"
        try:
            wb.save(filename)
            os.startfile(filename)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить файл: {str(e)}")

def main():
    root = tk.Tk()
    app = TestTimeApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()